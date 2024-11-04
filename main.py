
import cv2
from tensorflow.keras.models import model_from_json
import os
import numpy as np
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkthemes import ThemedStyle
from openpyxl import load_workbook
import face_recognition
import sys
import dlib
from imutils.video import VideoStream
import shutil
from imutils import face_utils
from tkinter.font import Font
from datetime import datetime, timedelta
import os
import tkinter as tk
from tkinter import ttk, filedialog
from ttkthemes import ThemedStyle
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
import openpyxl
def tontai(file_path):
    return os.path.isfile(file_path)
current_date = datetime.now()
first_day_of_current_month = current_date.replace(day=1)
last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
thangtruoc = last_day_of_previous_month.strftime("%m%Y")
current_date = datetime.now()
thangnay = current_date.strftime("%m%Y")
def eye_aspect_ratio(eye):
    A = np.linalg.norm(eye[1] - eye[5])
    B = np.linalg.norm(eye[2] - eye[4])
    C = np.linalg.norm(eye[0] - eye[3])
    ear = (A + B) / (2.0 * C)
    return ear
def copy_sheet(source_file, source_sheet_name, target_file, target_sheet_name):
    source_wb = openpyxl.load_workbook(source_file)
    if source_sheet_name in source_wb.sheetnames:
        source_sheet = source_wb[source_sheet_name]
        try:
            target_wb = openpyxl.load_workbook(target_file)
        except FileNotFoundError:
            target_wb = openpyxl.Workbook()
        if target_sheet_name in target_wb.sheetnames:
            target_wb.remove(target_wb[target_sheet_name])
        target_sheet = target_wb.create_sheet(title=target_sheet_name)
        for row in source_sheet.iter_rows(values_only=True):
            target_sheet.append(row)
        target_wb.save(target_file)
        print(f"Sheet '{source_sheet_name}' copied to '{target_sheet_name}' in '{target_file}'")
    else:
        print(f"Source sheet '{source_sheet_name}' not found in '{source_file}'")
file="TEST_"+str(thangnay)+".xlsx"
tt="TEST_"+str(thangtruoc)+".xlsx"
def copy_file(source_path, destination_path):
    try:
        shutil.copy2(source_path, destination_path)
        print(f"File {source_path} đã được sao chép thành công đến {destination_path}")
    except Exception as e:
        print(f"Lỗi khi sao chép file: {e}")
if tontai(file)==False and tontai(thangtruoc) == False:
    copy_file("TEST.xlsx",file)
if tontai(file)==False and tontai(thangtruoc)==True:
    copy_file("TEST.xlsx",file)
    copy_sheet(tt,"DATA",file,"DATA")
def is_first_day_of_month():
    today = datetime.now().day
    return today == 1
def is_excel_file_exists(file_path):
    return os.path.isfile(file_path) and file_path.lower().endswith('.xlsx')
class FaceRecognitionApp:
    def __init__(self, video_source=0, model_dir='antispoofing_models', employee_dir='NHANVIEN'):
        self.video = VideoStream(src=video_source).start()
        self.face_cascade = cv2.CascadeClassifier("models/haarcascade_frontalface_default.xml")
        self.fake = 0
        self.check = 0
        self.blink_counter = 0 
        self.tennv = ''
        self.known_images = []
        self.known_names = []
        self.detector = dlib.get_frontal_face_detector()
        self.predictor = dlib.shape_predictor("models/shape_predictor_68_face_landmarks.dat")
        self.load_models(model_dir)
        self.load_known_faces(employee_dir)
        self.attendance_history = []
        self.is_running = True 
    def load_models(self, model_dir):
        json_file = open(os.path.join(model_dir, 'antispoofing_model.json'), 'r')
        loaded_model_json = json_file.read()
        json_file.close()
        self.model = model_from_json(loaded_model_json)
        self.model.load_weights(os.path.join(model_dir, 'antispoofing_model.h5'))
    def load_known_faces(self, employee_dir):
        valid_image_extensions = ['.jpg', '.jpeg', '.png'] 
        for filename in os.listdir(employee_dir):
            path = os.path.join(employee_dir, filename)
            if any(path.lower().endswith(ext) for ext in valid_image_extensions):
                try:
                    image = face_recognition.load_image_file(path)
                    face_encodings = face_recognition.face_encodings(image)
                    if len(face_encodings) > 0:
                        encoding = face_encodings[0]
                        self.known_images.append(encoding)
                        self.known_names.append(filename)
                    else:
                        print(f"No face detected in {filename}")
                except Exception as e:
                    print(f"Error processing {filename}: {e}")
            else:
                print(f"Skipping non-image file: {filename}")
    def tim(self,file_path, employee_name, data):
        def chenhlech(time_str1, time_str2):
            time_format = "%H:%M"
            time1 = datetime.strptime(time_str1, time_format).time()
            time2=datetime.strptime(time_str2, time_format).time()
            time_difference = datetime.combine(datetime.min, time1) - datetime.combine(datetime.min, time2)
            if time_difference.total_seconds() < 0:
                time_difference = -time_difference
            result_time_str = str(time_difference)
            return result_time_str
        workbook = load_workbook(file_path)
        data_sheet = workbook["DATA"]
        today_date_str = datetime.now().date().strftime("%d")
        sheet_name_str = "Sheet" + today_date_str
        for row in range(1, data_sheet.max_row + 1):
            cell_value = data_sheet.cell(row=row, column=1).value
            if cell_value == employee_name:
                if sheet_name_str not in workbook.sheetnames:
                    sheet_name = workbook.create_sheet(title=sheet_name_str)
                else:
                    sheet_name = workbook[sheet_name_str]
                entry = sheet_name.max_row + 1
                gio=data_sheet.cell(row=row, column=4).value
                t=False
                for k in range(1,entry+1):
                    if sheet_name["A"+str(k)].value==employee_name:
                        cell = sheet_name.cell(row=k, column=5, value=employee_name)
                        t=True
                if t==False:
                    sheet_name["A"+str(entry)].value= employee_name
                    sheet_name["B"+str(entry)].value=data_sheet.cell(row=row, column=2).value
                    sheet_name["D"+str(entry)].value=chenhlech(data,gio)
                    if sheet_name['C' + str(entry)].value is None:
                        cell = sheet_name.cell(row=entry, column=3, value=employee_name)
                    else:
                        cell = sheet_name.cell(row=entry, column=5, value=employee_name)   
                cell.value = data
                workbook.save(file_path)
                return True
        print(f"Employee {employee_name} not found in sheet 'DATA'.")
        return False
    def recognize_person(self, frame):
        face_encodings = face_recognition.face_encodings(frame)
        if not face_encodings:
            return False
        unknown_face_encoding = face_encodings[0]
        results = face_recognition.compare_faces(self.known_images, unknown_face_encoding, tolerance=0.4)
        if any(results):
            index = results.index(True)
            file_name_with_extension = self.known_names[index]
            self.tennv = os.path.splitext(file_name_with_extension)[0]
            return True
        return False
    def process_frame(self, frame):
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            face = frame[y-5:y+h+5, x-5:x+w+5]
            resized_face = cv2.resize(face, (160, 160))
            resized_face = resized_face.astype("float") / 255.0
            resized_face = np.expand_dims(resized_face, axis=0)
            preds = self.model.predict(resized_face)[0]
            shape = self.predictor(gray, dlib.rectangle(x, y, x+w, y+h))
            shape = face_utils.shape_to_np(shape)
            left_eye = shape[36:42]
            right_eye = shape[42:48]
            left_ear = eye_aspect_ratio(left_eye)
            right_ear = eye_aspect_ratio(right_eye)
            average_ear = (left_ear + right_ear) / 2.0
            if average_ear < 0.2:
                self.blink_counter += 1
            if preds > 0.0007: 
                label = 'DUNG YEN'
                self.fake += 1
                self.check = 0
                cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                self.blink_counter=0
                if self.fake > 10:
                    cv2.putText(frame, 'CANH BAO: KHONG GIA MAO!', (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    self.fake = 0
            if preds < 0.0007 and self.blink_counter >= 1:
                self.check += 1
                label = "Dang xac thuc:" + str((self.check/7)*100)+'%'
                cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                if self.check > 7:
                    if self.recognize_person(frame):
                        current_time = datetime.now()
                        data_to_write = f"{current_time.hour:02d}:{current_time.minute:02d}"
                        self.tim(file, self.tennv, data_to_write)
                        message = f"Nhân viên được xác định: {self.tennv} lúc {data_to_write}"
                        messagebox.showinfo("Xác định nhân viên", message)
                    else:
                        messagebox.showinfo("Thông báo", "Không tìm thấy khuôn mặt trong tệp NHANVIEN.")
                    self.check = 0
                    self.blink_counter = 0
        return frame
    def run(self):
        while self.is_running:  
            try:
                frame = self.video.read()
                processed_frame = self.process_frame(frame)
                cv2.imshow('frame', processed_frame)
                key = cv2.waitKey(1)
                if key == 27: 
                    self.is_running = False 
            except Exception as e:
                print(f"Error: {e}")
        self.video.stop()
        cv2.destroyAllWindows()

import time
job_titles = ["ĐẦU BẾP", "TẠP VỤ", "QUẢN LÝ", "BẢO VỆ", "THU NGÂN"]
def slide(tab_frame, current_tab, target_tab):
    print(f"Current tab names: {tab_frame.tab_names()}")
    target_index = tab_frame.index(target_tab)
    current_index = tab_frame.index(current_tab)
    direction = 1 if target_index > current_index else -1
    for i in range(current_index, target_index, direction):
        tab_frame.forget(i)
        tab_frame.add(i + 1, text=f'Tab {i + 2}')  
        root.update_idletasks()
        root.update()
        root.after(10)
    print(f"Updated tab names: {tab_frame.tab_names()}")

def create_label_entry(parent, row, column, label_text, entry_var=None, entry_width=None, button_command=None):
    label = ttk.Label(parent, text=label_text)
    label.grid(row=row, column=column, pady=10, sticky='w')
    if entry_var is None:
        entry = ttk.Entry(parent)
    else:
        entry = ttk.Entry(parent, textvariable=entry_var)
    if entry_width:
        entry.config(width=entry_width)
    entry.grid(row=row, column=column + 1, pady=10, sticky='w')
    if button_command:
        button = ttk.Button(parent, text="Browse", command=button_command)
        button.grid(row=row, column=column + 2, pady=10, sticky='w')
    return entry
def show_add_employee_info():
    name_label.grid(row=1, column=0, pady=10)
    name_entry.grid(row=1, column=1, pady=10)
    position_label.grid(row=2, column=0, pady=10)
    position_combobox.grid(row=2, column=1, pady=10)
    hours_label.grid(row=3, column=0, pady=10)
    hours_entry.grid(row=3, column=1, pady=10)
    image_label.grid(row=4, column=0, pady=10)
    image_entry.grid(row=4, column=1, pady=10)
    browse_button.grid(row=4, column=2, pady=10)
    check_in_label.grid(row=5, column=0, pady=10)
    check_in_entry.grid(row=5, column=1, pady=10)
    salary_label.grid(row=6, column=0, pady=10)
    salary_entry.grid(row=6, column=1, pady=10)
    fixed_salary_label.grid(row=7, column=0, pady=10)
    fixed_salary_entry.grid(row=7, column=1, pady=10)
    revenue_percentage_label.grid(row=8, column=0, pady=10)
    revenue_percentage_entry.grid(row=8, column=1, pady=10)
    dedicated_salary_label.grid(row=9, column=0, pady=10)
    dedicated_salary_entry.grid(row=9, column=1, pady=10)
    add_button.grid(row=10, column=0, columnspan=3, pady=20)
    portrait_label.grid(row=0, column=3, rowspan=9, padx=20)
    update_button.grid_forget()
    portrait_label.grid_forget()
    attendance_history_button.grid_forget()
    slide(tab_frame, "Add Tab", "Update Tab")
def show_update_employee_info():
    name_label.grid(row=1, column=0, pady=10)
    name_entry.grid(row=1, column=1, pady=10)
    position_label.grid(row=2, column=0, pady=10)
    position_combobox.grid(row=2, column=1, pady=10)
    hours_label.grid(row=3, column=0, pady=10)
    hours_entry.grid(row=3, column=1, pady=10)
    check_in_label.grid(row=4, column=0, pady=10)
    check_in_entry.grid(row=4, column=1, pady=10)
    salary_label.grid(row=5, column=0, pady=10)
    salary_entry.grid(row=5, column=1, pady=10)
    fixed_salary_label.grid(row=6, column=0, pady=10)
    fixed_salary_entry.grid(row=6, column=1, pady=10)
    revenue_percentage_label.grid(row=7, column=0, pady=10)
    revenue_percentage_entry.grid(row=7, column=1, pady=10)
    dedicated_salary_label.grid(row=8, column=0, pady=10)
    dedicated_salary_entry.grid(row=8, column=1, pady=10)
    update_button.grid(row=9, column=0, columnspan=3, pady=20)
    add_button.grid_forget()
    image_label.grid_forget()
    image_entry.grid_forget()
    browse_button.grid_forget()
    attendance_history_button.grid_forget() 
    slide(tab_frame, "Update Tab", "Add Tab")
def add_employee():
    name = name_entry.get().strip()
    position = position_combobox.get()
    work_hours = hours_entry.get()
    image_path = image_entry.get()
    check_in_time = check_in_entry.get()
    salary = salary_entry.get()
    fixed_salary = fixed_salary_entry.get()
    revenue_percentage = revenue_percentage_entry.get()
    dedicated_salary = dedicated_salary_entry.get()
    if not all((name, position, work_hours, image_path, check_in_time, salary, fixed_salary, revenue_percentage, dedicated_salary)):
        messagebox.showinfo("Thông báo", "HÃY ĐIỀN ĐẦY ĐỦ THÔNG TIN.")
        return
    if not os.path.isfile(image_path):
        messagebox.showinfo("Thông báo", f"KHÔNG TÌM THẤY ẢNH {image_path}")
        return
    excel_path = file
    try:
        wb = load_workbook(excel_path)
        sheet = wb["DATA"]
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=position)
        sheet.cell(row=row, column=3, value=work_hours)
        sheet.cell(row=row, column=4, value=check_in_time)
        sheet.cell(row=row, column=5, value=salary)
        sheet.cell(row=row, column=6, value=fixed_salary)
        sheet.cell(row=row, column=7, value=revenue_percentage)
        sheet.cell(row=row, column=8, value=dedicated_salary)
        save_path = os.path.join("NHANVIEN", f"{name}.jpg")
        os.makedirs("NHANVIEN", exist_ok=True)
        shutil.copy2(image_path, save_path)
        wb.save(excel_path)
        messagebox.showinfo("Thông báo",f"NHÂN SỰ {name} ĐÃ ĐƯỢC CẬP NHẬT VÀO HỆ THỐNG.")
    except Exception as e:
        print(f"Error: {e}")
def update_employee():
    name = name_entry.get().strip()
    position = position_combobox.get()
    work_hours = hours_entry.get()
    check_in_time = check_in_entry.get()
    salary = salary_entry.get()
    fixed_salary = fixed_salary_entry.get()
    revenue_percentage = revenue_percentage_entry.get()
    dedicated_salary = dedicated_salary_entry.get()
    if not all((name, position, work_hours, check_in_time, salary, fixed_salary, revenue_percentage, dedicated_salary)):
        messagebox.showinfo("Thông báo", "HÃY ĐIỀN ĐẦY ĐỦ THÔNG TIN")
        return
    excel_path = file
    try:
        wb = load_workbook(excel_path)
        sheet = wb["DATA"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == name:
                sheet.cell(row=row, column=2, value=position)
                sheet.cell(row=row, column=3, value=work_hours)
                sheet.cell(row=row, column=4, value=check_in_time)
                sheet.cell(row=row, column=5, value=salary)
                sheet.cell(row=row, column=6, value=fixed_salary)
                sheet.cell(row=row, column=7, value=revenue_percentage)
                sheet.cell(row=row, column=8, value=dedicated_salary)
                break
        else:
            print(f"NHÂN SỰ {name} KHÔNG TỒN TẠI.")
            return
        wb.save(excel_path)
        print(f"NHÂN SỰ {name} CẬP NHẬT THÀNH CÔNG.")
    except Exception as e:
        print(f"Error: {e}")
face_recognition_app = FaceRecognitionApp()
def clock_in():
    current_time = datetime.now().strftime("%H:%M:%S")
    print(f"Đã chấm công vào lúc {current_time}")
def run_face_recognition():
    app = FaceRecognitionApp()
    result_tennv, result_time = app.run()
def on_closing():
    face_recognition_app.video.stop()  # Release the camera
    root.destroy()
    cv2.destroyAllWindows()
    sys.exit(0)
root = tk.Tk()
root.title("QUẢN LÝ NHÂN SỰ")
style = ThemedStyle(root)
style.set_theme("equilux")
label_font = Font(family="Helvetica", size=12, weight="bold")
tab_frame = ttk.Notebook(root)
tab_frame.grid(row=1, column=0, sticky="nsew")
content_frame = ttk.Frame(root, padding="20")
content_frame.grid(row=0, column=0, sticky="nsew")
add_employee_button = ttk.Button(content_frame, text="THÊM NHÂN VIÊN", command=show_add_employee_info)
add_employee_button.grid(row=0, column=0, pady=10)
update_employee_button = ttk.Button(content_frame, text="SỬA THÔNG TIN NHÂN VIÊN", command=show_update_employee_info)
update_employee_button.grid(row=0, column=1, pady=10)
clock_in_button = ttk.Button(content_frame, text="CHẤM CÔNG", command=run_face_recognition)
clock_in_button.grid(row=0, column=2, pady=10)
name_label = tk.Label(content_frame, text="HỌ VÀ TÊN:", font=label_font, borderwidth=2, relief="solid")
name_entry = ttk.Entry(content_frame, font=label_font)
position_label = tk.Label(content_frame, text="CHỨC VỤ:", borderwidth=2, relief="solid")
position_combobox = ttk.Combobox(content_frame, values=job_titles)
hours_label = tk.Label(content_frame, text="GIỜ ĐĂNG KÍ:", borderwidth=2, relief="solid")
hours_entry = ttk.Entry(content_frame)
image_label = tk.Label(content_frame, text="ẢNH CHÂN DUNG:", borderwidth=2, relief="solid")
image_entry = ttk.Entry(content_frame)
browse_button = ttk.Button(content_frame, text="Browse", command=lambda: image_entry.insert(0, filedialog.askopenfilename()))
check_in_label = tk.Label(content_frame, text="GIỜ VÀO CA:", borderwidth=2, relief="solid")
check_in_entry = ttk.Entry(content_frame)
salary_label = tk.Label(content_frame, text="SỐ CA ĐĂNG KÍ:", borderwidth=2, relief="solid")
salary_entry = ttk.Entry(content_frame, font=label_font)
add_button = ttk.Button(content_frame, text="THÊM NHÂN VIÊN", command=add_employee)
update_button = ttk.Button(content_frame, text="CẬP NHẬT THÔNG TIN NHÂN VIÊN", command=update_employee)
fixed_salary_label = tk.Label(content_frame, text="Lương cố định:", font=label_font, borderwidth=2, relief="solid")
fixed_salary_entry = ttk.Entry(content_frame)
revenue_percentage_label = tk.Label(content_frame, text="Phần trăm doanh thu:", borderwidth=2, relief="solid")
revenue_percentage_entry = ttk.Entry(content_frame)
dedicated_salary_label = tk.Label(content_frame, text="Lương theo sản phẩm:", borderwidth=2, relief="solid")
dedicated_salary_entry = ttk.Entry(content_frame)
portrait_label = ttk.Label(content_frame, text="Ảnh chân dung:", font=label_font)
portrait_label.grid(row=0, column=3, rowspan=9, padx=20)
portrait_label.grid_forget()
face_recognition_app = FaceRecognitionApp()
for i in range(8):
    content_frame.grid_rowconfigure(i, weight=1)
    content_frame.grid_columnconfigure(i, weight=1)
def show_attendance_history():
    try:
        file_path = 'TEST.xlsx'
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Sheet1'] 

        attended_employees = set()
        attendance_history = []
        for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True): 
            employee_name, check_in_time = row[0], row[4]
            if employee_name and check_in_time:
                attended_employees.add(employee_name)
                entry = f"{employee_name} - Giờ chấm công: {check_in_time}"
                attendance_history.append(entry)
        if attendance_history:
            history_message = "\n".join(attendance_history)
            messagebox.showinfo("Lịch sử chấm công", history_message)
        else:
            messagebox.showinfo("Lịch sử chấm công", "Không có dữ liệu chấm công.")
    except Exception as e:
        print(f"Error: {e}")
def hide_portrait_label():
    portrait_label.grid_forget()
attendance_history_button = ttk.Button(content_frame, text="Lịch sử chấm công", command=show_attendance_history)
attendance_history_button.grid(row=9, column=0, columnspan=3, pady=10)
root.protocol("WM_DELETE_WINDOW", on_closing) 
root.mainloop()