
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, MessageHandler, Filters,ConversationHandler
import os
import warnings
import openpyxl
import calendar
from datetime import datetime
TOKEN = "6987355245:AAFca599y7_i1qnsxS4ucVJuidTCWNDCShg"
PRINT_TONGKET, GET_REVENUE,xin,now,edit = range(5)
AUTHORIZED_USERS_FILE = "quanly.txt"
AUTHORIZED_USERS = set()
def load_authorized_users():
    try:
        with open(AUTHORIZED_USERS_FILE, "r") as file:
            lines = file.readlines()
            return set(str(line.strip()) for line in lines)
    except FileNotFoundError:
        return set()
AUTHORIZED_USERS = load_authorized_users()
print(AUTHORIZED_USERS)
def start(update, context):
    user_id= update.effective_user.username
    chu = "DuyQuaan"
    update.message.reply_text("Xin chào "+ user_id) 
    if user_id == chu:
        keyboard = [
        [InlineKeyboardButton("In Tổng Kết", callback_data='print_tongket')],
        [InlineKeyboardButton("NHÂN VIÊN HIỆN CÓ TẠI QUÁN", callback_data='nhan_vien_tai_quan')],
        [InlineKeyboardButton("XIN ĐẾN MUỘN", callback_data='xin_nghi')],
        [InlineKeyboardButton("THỐNG KÊ", callback_data='now')],
        [InlineKeyboardButton("Chỉnh sửa Quản Lý", callback_data='edit')],
    ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        update.message.reply_text('Chọn một hành động:', reply_markup=reply_markup)
        return PRINT_TONGKET
    if user_id != chu and user_id not in AUTHORIZED_USERS :
        update.message.reply_text("Xin lỗi, bạn không có quyền sử dụng bot của GASACH.")
        return ConversationHandler.END
    if user_id in AUTHORIZED_USERS:
        keyboard = [
        [InlineKeyboardButton("In Tổng Kết", callback_data='print_tongket')],
        [InlineKeyboardButton("NHÂN VIÊN HIỆN CÓ TẠI QUÁN", callback_data='nhan_vien_tai_quan')],
        [InlineKeyboardButton("XIN NGHỈ", callback_data='xin_nghi')],
        [InlineKeyboardButton("THỐNG KÊ", callback_data='now')],
    ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        update.message.reply_text('Chọn một hành động:', reply_markup=reply_markup)
        return PRINT_TONGKET
def stop(update, context):
    user_id = update.effective_user.id
    if user_id not in AUTHORIZED_USERS:
        update.message.reply_text("Xin lỗi, bạn không có quyền sử dụng lệnh tắt bot.")
        return
    update.message.reply_text("Bot đã tắt.")
    updater = context.bot_data.get('updater')
    if updater:
        updater.stop()
current_date = datetime.now()
thangnay = current_date.strftime("%m%Y")
file="TEST_"+str(thangnay)+".xlsx"
def chenhlech(time_str1, time_str2):
    time_format = "%H:%M"
    time1 = datetime.strptime(time_str1, time_format).time()
    time2 = datetime.strptime(time_str2, time_format).time()
    time_difference = datetime.combine(datetime.min, time1) - datetime.combine(datetime.min, time2)
    if time_difference.total_seconds() < 0:
        time_difference = -time_difference
        result_time_str = str(time_difference)
    return result_time_str
def button_click(update, context):
    query = update.callback_query
    button_data = query.data
    if button_data == 'print_tongket':
        query.edit_message_text(text='Nhập doanh thu và tháng muốn tổng kết cách nhau dấu :')
        return GET_REVENUE
    elif button_data == 'nhan_vien_tai_quan':
        gio = datetime.now().date().strftime("%d")
        wb = openpyxl.load_workbook(file)
        sheet_name = "Sheet" + str(gio)
        sua = wb[sheet_name]
        employees_at_quan = []
        vitri = []
        for i in range(2, sua.max_row + 1):
            t = sua.cell(row=i, column=3).value
            if t is None:
                employees_at_quan.append(str(sua.cell(row=i, column=1).value))
                vitri.append(str(sua.cell(row=i, column=2).value))
            elif chenhlech(sua.cell(row=i, column=3).value, str(sua.cell(row=i, column=5).value)) < "00:05":
                employees_at_quan.append(str(sua.cell(row=i, column=1).value))
                vitri.append(str(sua.cell(row=i, column=2).value))
        if employees_at_quan:
            for i in range(len(employees_at_quan)):
                employees_at_quan[i] += "(" + vitri[i] + ")"
            message_text = f'Nhân viên tại quán: {", ".join(employees_at_quan)} '
        else:
            message_text = 'Hiện không có nhân viên tại quán.'
        context.bot.send_message(chat_id=query.message.chat_id, text=message_text)
    elif button_data == 'xin_nghi':
        query.edit_message_text(text='Nhập thông tin xin nghỉ của nhân viên và ngày xin nghỉ cách nhau dấu :')
        return xin
    elif button_data == 'now':
        query.edit_message_text(text='Nhập tháng muốn lấy chi tiết:')
        return now
    elif button_data == 'edit':
        query.edit_message_text("Xoá: id nếu muốn xoá, Thêm: id nếu muốn thêm quản lý:")
        return edit
def edit(update, context):
    revenue = update.message.text
    context.user_data['revenue'] = revenue
    words = revenue.split(":")
    if len(words) >= 2:
        command = words[0]
        user_id = words[1]
        if command == "Thêm":
            with open(AUTHORIZED_USERS_FILE, "a") as file:
                file.write(str(user_id) + "\n")
            update.message.reply_text(f"Đã thêm {user_id} vào hệ thống.")
        elif command == "Xoá":
            try:
                with open(AUTHORIZED_USERS_FILE, "r") as file:
                    lines = file.readlines()
                with open(AUTHORIZED_USERS_FILE, "w") as file:
                    for line in lines:
                        if line.strip() != str(user_id):
                            file.write(line)
                update.message.reply_text(f"Đã xoá {user_id} khỏi hệ thống.")
            except FileNotFoundError:
                update.message.reply_text("Không tìm thấy file quản lý.")
        else:
            update.message.reply_text("Lệnh không hợp lệ. Hãy sử dụng 'Thêm' hoặc 'Xoá'.")
    else:
        update.message.reply_text("Lệnh không hợp lệ. Hãy sử dụng 'Thêm' hoặc 'Xoá'.")
    return ConversationHandler.END
def tontai(file_path):
    return os.path.isfile(file_path)
def now(update, context):
    revenue = update.message.text
    context.user_data['revenue'] = revenue
    print(revenue)
    t=str(revenue)
    if len(t)==1:
        t="0"+t+str(current_date.strftime("%Y"))
    else:
        t=t+str(current_date.strftime("%Y"))
    document_path = "TEST_"+t + ".xlsx"
    if not os.path.isfile(document_path):
        update.message.reply_text(f"File {revenue} không có dữ liệu.")
        return ConversationHandler.END
    context.bot.send_document(
        chat_id=update.message.chat_id,
        document=open(document_path, "rb"),
        filename=os.path.basename(document_path),
        caption=f"Đây là Tổng Kết Tháng  {revenue}"
    )
    return ConversationHandler.END
def get_revenue(update, context):
    revenue = update.message.text
    context.user_data['revenue'] = revenue
    def get_current_month():
        current_date = datetime.now()
        current_month = current_date.month
        return current_month
    thang = get_current_month()
    def get_number_of_days_in_current_month():
        current_date = datetime.now()
        _, days_in_month = calendar.monthrange(current_date.year, current_date.month)
        return days_in_month
    phan_tach = revenue.split(":")
    doanhthu=phan_tach[0]
    if len(phan_tach)<2:
        update.message.reply_text(f"Kiểu dữ liệu sai.")
        return ConversationHandler.END
    if int(phan_tach[1])is None or int(phan_tach[0])is None:
        update.message.reply_text(f"Kiểu dữ liệu sai.")
        return ConversationHandler.END
    if len(phan_tach[1])==1:
        t= "TEST_0"+str(phan_tach[1])+str(datetime.now().strftime("%Y"))+".xlsx"
        phan_tach[1]=t
    if not os.path.isfile(phan_tach[1]):
        update.message.reply_text(f"File {phan_tach[1]} không có dữ liệu.")
        return ConversationHandler.END
    file=phan_tach[1]
    ngay = get_number_of_days_in_current_month()
    wb = openpyxl.load_workbook(file)
    sheet_name = "DATA"
    sheet = wb[sheet_name]
    tongket = openpyxl.load_workbook('TK.xlsx')
    if "TK" in tongket.sheetnames:
        tk = tongket["TK"]
    else:
        tk = tongket.create_sheet("TK")
    ten = {}
    for i in range(2, sheet.max_row + 1):
        tong=0
        index = tk.max_row + 1
        entry_cell = sheet.cell(row=i, column=1).value
        tk["G" + str(index)] = sheet.cell(row=i, column=8).value
        if entry_cell not in ten:
            ten[entry_cell] = 0
        tk["A" + str(index)] = entry_cell
        tk["B" + str(index)] = sheet.cell(row=i, column=2).value
        tk["C" + str(index)] = sheet.cell(row=i, column=6).value
        print(sheet.cell(row=i, column=5).value)
        print(sheet.cell(row=i, column=6).value)
        tong+=int(sheet.cell(row=i, column=6).value)
        for j in range(1, ngay + 1):
            sheet_name = "Sheet" + str(j)
            if sheet_name in wb.sheetnames:
                sh = wb[sheet_name]
                t = sh.max_row
                for k in range(1, t + 1):
                    if sh["A" + str(k)].value == entry_cell:
                        if sh["C" + str(k)].value != None or sh["C" + str(k)].value != None :
                            ten[entry_cell] += 1
        nghi=sheet.cell(row=i, column=5).value
        tk["E" + str(index)]= int(nghi)-ten[entry_cell]
        tk["D" + str(index)] = ten[entry_cell]
        t=(int((sheet.cell(row=i, column=7).value))*int(doanhthu))/100
        tk["F" + str(index)] = t
        tong+= t
        tk["H" + str(index)] = tong
    tongket.save("TONGKETTHANG" + str(thang) + ".xlsx")
    document_path = "TONGKETTHANG" + str(thang) + ".xlsx"
    context.bot.send_document(
    chat_id=update.message.chat_id,
    document=open(document_path, "rb"),
    filename=os.path.basename(document_path),
    caption=f"Đây là Tổng Kết Tháng {phan_tach[1]}. Doanh thu: {phan_tach[0]}"
    )
    return ConversationHandler.END
def xinnghi(update, context):
    revenue = update.message.text
    context.user_data['revenue'] = revenue
    gio = datetime.now().date().strftime("%d")
    phan_tach = revenue.split(":")
    if len(phan_tach)<2:
        update.message.reply_text(f"Nhập sai cú pháp.")
        return ConversationHandler.END
    wb = openpyxl.load_workbook(file)
    sheet_name = "Sheet"+str(phan_tach[1])
    sua = wb[sheet_name]
    for i in range(1, sua.max_row + 1):
        if revenue == sua.cell(row=i, column=1).value:
            sua.cell(row=i, column=4).value = 0
            message_text = f'Nhân viên: {phan_tach[0]} đã được cập nhật'
            break
    else:
        message_text = f'Nhân viên: {phan_tach[0]} chưa đến quán vào ngày {phan_tach[1]} tháng này'
    wb.save("TEST.xlsx")
    context.bot.send_message(chat_id=update.effective_chat.id, text=message_text)
    return ConversationHandler.END
def main():
    warnings.filterwarnings("ignore", category=UserWarning)
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher
    job_queue = updater.job_queue
    dp.add_handler(CommandHandler('start', start))

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            PRINT_TONGKET: [CallbackQueryHandler(button_click)],
            GET_REVENUE: [MessageHandler(Filters.text & ~Filters.command, get_revenue)],
            xin: [MessageHandler(Filters.text & ~Filters.command, xinnghi)],
            now: [MessageHandler(Filters.text & ~Filters.command, now)],
            edit: [MessageHandler(Filters.text & ~Filters.command, edit)],
        },
        fallbacks=[],
    )
    dp.add_handler(conv_handler)
    dp.add_handler(CommandHandler('stop', stop))
    try:
        updater.start_polling()
        updater.idle()
    except telegram.error.Conflict as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
if __name__ == '__main__':
    main()

